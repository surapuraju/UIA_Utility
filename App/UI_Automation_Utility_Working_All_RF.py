#-------------------------------------------------------------------------------
# Name:        UI Automation Utility
# Purpose:     UI Automation Utility
#
# Author:      Raju Surapuraju
#
# Created:     08-Mar-2025
# Copyright:   (c) Raju Surapuraju 2025
# Licence:     Raju Surapuraju
#-------------------------------------------------------------------------------
import os
import sys
import json
import time
import cv2
import pyautogui
import webbrowser
import configparser
from openpyxl import load_workbook

# Determine Base Directory (Handles PyInstaller Paths)
def get_base_path():
    """Determine base directory whether running as script or .exe."""
    if getattr(sys, 'frozen', False):  # Running as .exe
        base_path = os.path.dirname(sys.executable)
    else:  # Running as script
        base_path = os.path.abspath(os.path.dirname(__file__))

    return os.path.dirname(base_path)

BASE_DIR = get_base_path()

# Paths to Config and JSON Files
CONFIG_PATH = os.path.join(BASE_DIR, "Config", "configFile.ini")
JSON_PATH = os.path.join(BASE_DIR, "Config", "Properties_1.JSON")
EXCEL_PATH = None  # To be set dynamically

RUNTIME_SCREEN_DIR = os.path.join(BASE_DIR, "RunTime")
os.makedirs(RUNTIME_SCREEN_DIR, exist_ok=True)

# Load Config File with Debugging
config = configparser.ConfigParser()
if not os.path.exists(CONFIG_PATH):
    raise FileNotFoundError(f"❌ Config file missing: {CONFIG_PATH}")

config.read(CONFIG_PATH, encoding="utf-8")

def get_config_value(section, key, default=None):
    """Fetch config value with debugging."""
    try:
        value = config[section][key]
        print(f"🔹 Loaded [{section}] -> {key}: {value}")
        return value
    except KeyError:
        print(f"⚠️ Warning: Missing key [{section}] -> {key}. Using default: {default}")
        return default

# Read values from config.ini
url = get_config_value("DEFAULT", "url")
usernameConfig = get_config_value("DEFAULT", "username", "default_user")
passwordConfig = get_config_value("DEFAULT", "password", "default_pass")
test_records_to_create = get_config_value("DEFAULT", "test_records_to_create", "1")
timeout = int(get_config_value("Settings", "timeout", 10))
confidence = float(get_config_value("Settings", "confidence", 0.8))
excel_file_name = get_config_value("DEFAULT", "excel_file_input")

if not url:
    raise KeyError("❌ Critical: 'url' key is missing in configFile.ini")

EXCEL_PATH = os.path.join(BASE_DIR, "Data", excel_file_name)

print(f"🔹 Base Directory: {BASE_DIR}")
print(f"🔹 Config Path: {CONFIG_PATH}")
print(f"🔹 JSON Path: {JSON_PATH}")
print(f"🔹 Excel File Path: {EXCEL_PATH}")

# Function to Read JSON File
def read_json_file(json_path):
    """Reads a JSON file and returns the data as a list of dictionaries."""
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"⚠️ JSON file missing: {json_path}")

    try:
        with open(json_path, "r", encoding="utf-8") as file:
            return json.load(file)
    except json.JSONDecodeError as e:
        raise ValueError(f"⚠️ Error decoding JSON: {e}")

actions_array = read_json_file(JSON_PATH)

# Function to Read Excel Data Using openpyxl
def read_excel_data(excel_path):
    """Reads data from an Excel file (openpyxl) and returns it as a list of dictionaries."""
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"⚠️ Excel file not found: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]  # First row as header
    data_list = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_list.append(dict(zip(headers, row)))  # Convert row to dictionary

    wb.close()
    return data_list

# Read Excel Data
try:
    account_records = read_excel_data(EXCEL_PATH)
    print("✅ Excel file loaded successfully!")
except FileNotFoundError as e:
    print(e)
    sys.exit(1)

# Function to Launch Web App
def launchApp():
    """Opens the browser and navigates to the web application"""
    webbrowser.open(url)
    time.sleep(5)

# Function to Capture Screenshot
def screenCapture():
    """Captures a screenshot and saves it for processing"""
    screenshot_path = os.path.join(RUNTIME_SCREEN_DIR, "process_screen.png")
    screenshot = pyautogui.screenshot()
    screenshot.save(screenshot_path)
    return cv2.imread(screenshot_path)

# Function to Find and Click Elements
def find_and_click(screen, template_path, threshold=0.8):
    """Finds an element using OpenCV template matching and clicks it if found"""
    if not os.path.exists(template_path):
        print(f"⚠️ Image not found: {template_path}")
        return False

    template = cv2.imread(template_path, cv2.IMREAD_GRAYSCALE)
    screen_gray = cv2.cvtColor(screen, cv2.COLOR_BGR2GRAY)

    result = cv2.matchTemplate(screen_gray, template, cv2.TM_CCOEFF_NORMED)
    _, max_val, _, max_loc = cv2.minMaxLoc(result)

    if max_val >= threshold:
        x, y = max_loc
        h, w = template.shape
        center_x, center_y = x + w // 2, y + h // 2

        pyautogui.click(center_x, center_y)
        time.sleep(1)
        return True
    else:
        print(f"⚠️ Could not find {template_path} (Confidence: {max_val})")
        return False

def setText_action(objectId, data, screen):
    image_path = os.path.join(BASE_DIR, "Objects", objectId)
    if find_and_click(screen, image_path):
        data = str(data)  # Convert to string to avoid TypeError
        pyautogui.write(data, interval=0.1)
        print(f"✅ Entered: {data}")

# Function to Perform Click Action
def click_action(objectId, screen):
    image_path = os.path.join(BASE_DIR, "Objects", objectId)
    find_and_click(screen, image_path)

# 🚀 Start Automation Process
for record in account_records:
    launchApp()
    #print("RECORD:", record)

    # Process JSON Actions
    for action in actions_array:
        field_name = action.get("field_name")
        action_type = action.get("action")
        objectId = action.get("objectId")

        # Fetch corresponding data from the Excel file
        data = record.get(field_name, "")
        #print("RECORD DATA:", data)

        if field_name == "Username":
            data = usernameConfig
        elif field_name == "Password":
            data = passwordConfig

        if not objectId:
            print("⚠️ Skipping action due to missing 'objectId'.")
            continue

        # Take a new screenshot before every action
        screen = screenCapture()

        if action_type == "setText":
            setText_action(objectId, data, screen)
        elif action_type == "Click":
            click_action(objectId, screen)

print("🎉 Automation process completed successfully!")
